import datetime
import os

from openpyxl import Workbook, load_workbook

from config.cfg import xlsx_path


class XlsxWriter():

    def __init__(self, filename, dirpath):
        self.filepath = os.path.join(dirpath, filename)
        try:
            self.workbook = load_workbook(self.filepath)
        except Exception:
            self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.line = 0

    def create_sheet(self):
        ws = self.workbook.create_sheet()
        ws.title = datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        self.worksheet = ws
        print(self.worksheet)

    def new_line(self, *args):

        def set_cell(sheet, row, column, value):
            wc = sheet.cell(row, column)
            wc.value = value

        self.line += 1
        line = self.line
        ws = self.worksheet
        for i, value in enumerate(args):
            set_cell(ws, line, i + 1, value)

    def save(self):
        self.workbook.save(self.filepath)

    def close(self):
        self.save()


if __name__ == '__main__':
    xlsxwriter = XlsxWriter('a.xlsx', xlsx_path)
    xlsxwriter.create_sheet()
    xlsxwriter.new_line(1, 2, 2)
    xlsxwriter.new_line(2, 3, 3, 4)
    xlsxwriter.close()
